"""Excel 读写工具 — openpyxl 驱动。"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from pydantic import BaseModel, Field

from app.models.permission import PermissionLevel
from app.models.tool import ToolResult
from app.tools.base import BaseTool, ToolCategory, ToolUseContext


def _resolve_path(path: str) -> Path:
    p = Path(path).expanduser().resolve()
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


class ExcelOpenInput(BaseModel):
    path: str = Field(description="File path (creates new if not exists)")
    sheet_name: Optional[str] = Field(default=None, description="Sheet name (default: active sheet)")


class ExcelOpenTool(BaseTool):
    name = "excel_open"
    description = "Open or create an Excel file. Returns sheet names and dimensions."
    input_schema = ExcelOpenInput
    permission_level = PermissionLevel.READ
    category = ToolCategory.FILE

    async def execute(self, input_data: ExcelOpenInput, context: ToolUseContext | None = None) -> ToolResult:
        path = _resolve_path(input_data.path)
        from openpyxl import load_workbook, Workbook
        if path.exists():
            wb = load_workbook(path, data_only=True)
        else:
            wb = Workbook()
            wb.save(path)

        info = {
            "sheets": wb.sheetnames,
            "active": wb.active.title if wb.active else "",
        }
        output = f"[excel] Opened: {path.name}\nSheets: {', '.join(info['sheets'])}\nActive: {info['active']}"
        return ToolResult(tool_call_id="", output=output, metadata=info)


class ExcelReadInput(BaseModel):
    path: str
    range: str = Field(description="Cell range (e.g. 'A1:C10' or 'A1')")
    sheet_name: Optional[str] = None


class ExcelReadTool(BaseTool):
    name = "excel_read"
    description = "Read cell values from an Excel file."
    input_schema = ExcelReadInput
    permission_level = PermissionLevel.READ
    category = ToolCategory.FILE

    async def execute(self, input_data: ExcelReadInput, context: ToolUseContext | None = None) -> ToolResult:
        path = _resolve_path(input_data.path)
        if not path.exists():
            return ToolResult(tool_call_id="", output=f"[excel] File not found: {path}", is_error=True)

        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb[input_data.sheet_name] if input_data.sheet_name else wb.active

        from openpyxl.utils import range_boundaries
        try:
            min_col, min_row, max_col, max_row = range_boundaries(input_data.range)
        except Exception:
            cell = ws[input_data.range]
            return ToolResult(tool_call_id="", output=f"[excel] {input_data.range} = {cell.value}", metadata={"value": cell.value})

        data = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            data.append([str(v) if v is not None else "" for v in row])

        md_lines = ["| " + " | ".join(data[0]) + " |"] if data else []
        if md_lines:
            md_lines.append("| " + " | ".join(["---"] * len(data[0])) + " |")
        for row in data[1:]:
            md_lines.append("| " + " | ".join(row) + " |")

        output = f"[excel] Read {len(data)} rows × {len(data[0]) if data else 0} cols\n\n" + "\n".join(md_lines)
        return ToolResult(tool_call_id="", output=output[:30000], metadata={"rows": len(data), "data": data})


class ExcelWriteInput(BaseModel):
    path: str
    range: str = Field(description="Starting cell, e.g. 'A1'")
    values: list[list[Any]] = Field(description="2D array of values to write")
    sheet_name: Optional[str] = None


class ExcelWriteTool(BaseTool):
    name = "excel_write"
    description = "Write data to an Excel file starting at a given cell."
    input_schema = ExcelWriteInput
    permission_level = PermissionLevel.WRITE
    category = ToolCategory.FILE

    async def execute(self, input_data: ExcelWriteInput, context: ToolUseContext | None = None) -> ToolResult:
        path = _resolve_path(input_data.path)
        from openpyxl import load_workbook, Workbook
        wb = load_workbook(path) if path.exists() else Workbook()
        ws = wb[input_data.sheet_name] if input_data.sheet_name and input_data.sheet_name in wb.sheetnames else wb.active
        if input_data.sheet_name and input_data.sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(input_data.sheet_name)

        from openpyxl.utils import coordinate_to_tuple
        start_col, start_row = coordinate_to_tuple(input_data.range)
        for i, row_data in enumerate(input_data.values):
            for j, val in enumerate(row_data):
                cell = ws.cell(row=start_row + i, column=start_col + j)
                cell.value = val

        wb.save(path)
        return ToolResult(tool_call_id="", output=f"[excel] Written {len(input_data.values)} rows × {len(input_data.values[0]) if input_data.values else 0} cols to {path.name}")


class ExcelFormulaInput(BaseModel):
    path: str
    cell: str = Field(description="Target cell, e.g. 'D2'")
    formula: str = Field(description="Excel formula, e.g. '=SUM(B2:B10)'")
    sheet_name: Optional[str] = None


class ExcelFormulaTool(BaseTool):
    name = "excel_formula"
    description = "Apply an Excel formula to a cell."
    input_schema = ExcelFormulaInput
    permission_level = PermissionLevel.WRITE
    category = ToolCategory.FILE

    async def execute(self, input_data: ExcelFormulaInput, context: ToolUseContext | None = None) -> ToolResult:
        path = _resolve_path(input_data.path)
        if not path.exists():
            return ToolResult(tool_call_id="", output=f"[excel] File not found: {path}", is_error=True)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb[input_data.sheet_name] if input_data.sheet_name else wb.active
        ws[input_data.cell] = input_data.formula
        wb.save(path)
        return ToolResult(tool_call_id="", output=f"[excel] Formula set: {input_data.cell} = {input_data.formula}")


class ExcelChartInput(BaseModel):
    path: str
    chart_type: str = Field(description="Chart type: bar | line | pie | scatter")
    data_range: str = Field(description="Data range, e.g. 'A1:B10'")
    title: Optional[str] = None
    sheet_name: Optional[str] = None


class ExcelChartTool(BaseTool):
    name = "excel_chart"
    description = "Create a chart in an Excel file."
    input_schema = ExcelChartInput
    permission_level = PermissionLevel.WRITE
    category = ToolCategory.FILE

    async def execute(self, input_data: ExcelChartInput, context: ToolUseContext | None = None) -> ToolResult:
        path = _resolve_path(input_data.path)
        if not path.exists():
            return ToolResult(tool_call_id="", output=f"[excel] File not found: {path}", is_error=True)
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
        wb = load_workbook(path)
        ws = wb[input_data.sheet_name] if input_data.sheet_name else wb.active

        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(input_data.data_range)
        data = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

        chart_map = {"bar": BarChart, "line": LineChart, "pie": PieChart, "scatter": ScatterChart}
        chart_cls = chart_map.get(input_data.chart_type)
        if not chart_cls:
            return ToolResult(tool_call_id="", output=f"[excel] Unknown chart type: {input_data.chart_type}", is_error=True)

        chart = chart_cls()
        chart.add_data(data, titles_from_data=True)
        if input_data.title:
            chart.title = input_data.title
        ws.add_chart(chart, f"E{max_row + 2}")
        wb.save(path)
        return ToolResult(tool_call_id="", output=f"[excel] Created {input_data.chart_type} chart in {path.name}")


def register_excel_tools(registry) -> None:
    for t in [ExcelOpenTool, ExcelReadTool, ExcelWriteTool, ExcelFormulaTool, ExcelChartTool]:
        registry.register(t())
