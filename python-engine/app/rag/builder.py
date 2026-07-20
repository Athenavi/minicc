# RAG 构建器 — 支持双向量数据库 + 混合解析器 + LangChain 分块
from __future__ import annotations

import asyncio
import logging
import uuid
from typing import Optional, AsyncIterator
from enum import Enum

from app.config import settings
from app.gateway.router import GatewayRouter
from app.interfaces.llm import LLMProvider
from app.interfaces.vectorstore import VectorStore

logger = logging.getLogger(__name__)


class VectorDBType(str, Enum):
    """向量数据库类型"""
    MILVUS = "milvus"
    PGVECTOR = "pgvector"


class ParserType(str, Enum):
    """解析器类型"""
    UNSTRUCTURED = "unstructured"
    MARKITDOWN = "markitdown"
    LIGHTWEIGHT = "lightweight"


class RAGBuilder:
    """RAG 知识库构建器 — 支持多种配置"""

    def __init__(
        self,
        llm_gateway: GatewayRouter,
        vector_store: VectorStore | None = None,
        vector_db_type: VectorDBType = VectorDBType.MILVUS,
        parser_type: ParserType = ParserType.UNSTRUCTURED,
    ):
        self._gateway = llm_gateway
        self._vector_store = vector_store
        self._vector_db_type = vector_db_type
        self._parser_type = parser_type
        self._milvus_connected = False

    def _ensure_milvus(self):
        """确保 Milvus 连接"""
        if not self._milvus_connected:
            from pymilvus import connections
            host = settings.milvus_address.split(":")[0]
            port = int(settings.milvus_address.split(":")[1]) if ":" in settings.milvus_address else 19530
            connections.connect(alias="default", host=host, port=port)
            self._milvus_connected = True
            logger.info("Milvus connected: %s:%d", host, port)

    async def build_document(
        self,
        kb_id: str,
        doc_id: str,
        content: bytes,
        file_type: str,
        filename: str,
        tenant_id: str,
        vector_db: str | None = None,
        parser_type: str | None = None,
    ) -> AsyncIterator[dict]:
        """
        构建单个文档的 RAG 索引（SSE 流式返回进度）
        """
        # 使用指定的向量数据库类型，或默认
        db_type = VectorDBType(vector_db) if vector_db else self._vector_db_type
        parse_type = ParserType(parser_type) if parser_type else self._parser_type

        try:
            # Step 1: 解析
            yield {"type": "progress", "step": "parsing", "progress": 0.1}
            parse_result = self._parse_document(content, file_type, filename, parse_type)
            if parse_result.get("error"):
                yield {"type": "error", "message": parse_result["error"]}
                return

            text = parse_result["text"]
            char_count = parse_result["char_count"]
            if not text.strip():
                yield {"type": "error", "message": "文档内容为空"}
                return

            # Step 2: 分块
            yield {"type": "progress", "step": "chunking", "progress": 0.3}
            chunks = self._chunk_text(text)
            if not chunks:
                yield {"type": "error", "message": "分块结果为空"}
                return

            # Step 3: 嵌入
            yield {"type": "progress", "step": "embedding", "progress": 0.5}
            embeddings = await self._compute_embeddings(chunks)
            yield {"type": "progress", "step": "embedding", "progress": 0.8}

            # Step 4: 存储
            yield {"type": "progress", "step": "storing", "progress": 0.9}
            await self._store_vectors(
                kb_id, doc_id, tenant_id, chunks, embeddings, db_type
            )

            yield {"type": "progress", "step": "storing", "progress": 1.0}
            yield {
                "type": "complete",
                "chunk_count": len(chunks),
                "char_count": char_count,
                "page_count": parse_result.get("page_count", 0),
            }
        except Exception as e:
            logger.error("文档构建失败: %s", e)
            yield {"type": "error", "message": str(e)}

    def _parse_document(
        self,
        content: bytes,
        file_type: str,
        filename: str,
        parser_type: ParserType,
    ) -> dict:
        """解析文档"""
        if parser_type == ParserType.UNSTRUCTURED:
            return self._parse_with_unstructured(content, file_type, filename)
        elif parser_type == ParserType.MARKITDOWN:
            return self._parse_with_markitdown(content, file_type, filename)
        elif parser_type == ParserType.LIGHTWEIGHT:
            return self._parse_lightweight(content, file_type, filename)
        else:
            return {"error": f"不支持的解析器类型: {parser_type}"}

    def _parse_with_unstructured(
        self,
        content: bytes,
        file_type: str,
        filename: str,
    ) -> dict:
        """使用 Unstructured 解析"""
        try:
            from unstructured.partition.auto import partition

            # 使用通用 partition 函数自动处理所有文件类型
            import tempfile
            import os
            suffix = f".{file_type}" if file_type else ""
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            try:
                elements = partition(filename=tmp_path)
            finally:
                os.unlink(tmp_path)

            text = "\n\n".join([str(el) for el in elements])
            return {
                "text": text,
                "char_count": len(text),
                "page_count": len([el for el in elements if type(el).__name__ == "PageBreak"]),
            }
        except Exception as e:
            logger.warning("Unstructured 解析失败: %s", e)
            return {"error": str(e)}

    def _parse_with_markitdown(
        self,
        content: bytes,
        file_type: str,
        filename: str,
    ) -> dict:
        """使用 markitdown 解析"""
        try:
            from markitdown import MarkItDown

            md = MarkItDown()
            result = md.convert(content)
            return {
                "text": result.text_content,
                "char_count": len(result.text_content),
                "page_count": 1,
            }
        except Exception as e:
            logger.warning("markitdown 解析失败: %s", e)
            return {"error": str(e)}

    def _parse_lightweight(
        self,
        content: bytes,
        file_type: str,
        filename: str,
    ) -> dict:
        """轻量级解析"""
        try:
            if file_type == "pdf":
                return self._parse_pdf_lightweight(content)
            elif file_type in ("docx", "doc"):
                return self._parse_docx_lightweight(content)
            elif file_type in ("xlsx", "xls"):
                return self._parse_xlsx_lightweight(content)
            elif file_type == "html":
                return self._parse_html_lightweight(content)
            elif file_type == "markdown":
                return {"text": content.decode("utf-8"), "char_count": len(content), "page_count": 1}
            else:
                return {"error": f"轻量级解析器不支持: {file_type}"}
        except Exception as e:
            logger.warning("轻量级解析失败: %s", e)
            return {"error": str(e)}

    def _parse_pdf_lightweight(self, content: bytes) -> dict:
        """使用 PyMuPDF 解析 PDF"""
        import fitz
        import io

        doc = fitz.open(stream=content, filetype="pdf")
        try:
            text_parts = []
            for page in doc:
                text_parts.append(page.get_text())
            text = "\n\n".join(text_parts)
            return {
                "text": text,
                "char_count": len(text),
                "page_count": len(doc),
            }
        finally:
            doc.close()

    def _parse_docx_lightweight(self, content: bytes) -> dict:
        """使用 python-docx 解析 Word"""
        from docx import Document
        import io

        doc = Document(io.BytesIO(content))
        try:
            text = "\n\n".join([para.text for para in doc.paragraphs if para.text.strip()])
            return {
                "text": text,
                "char_count": len(text),
                "page_count": 1,
            }
        finally:
            doc.close()

    def _parse_xlsx_lightweight(self, content: bytes) -> dict:
        """使用 openpyxl 解析 Excel"""
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(content))
        try:
            text_parts = []
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        text_parts.append(row_text)
            text = "\n".join(text_parts)
            return {
                "text": text,
                "char_count": len(text),
                "page_count": len(wb.sheetnames),
            }
        finally:
            wb.close()

    def _parse_html_lightweight(self, content: bytes) -> dict:
        """使用 BeautifulSoup 解析 HTML"""
        from bs4 import BeautifulSoup

        soup = BeautifulSoup(content, "html.parser")
        text = soup.get_text(separator="\n", strip=True)
        return {
            "text": text,
            "char_count": len(text),
            "page_count": 1,
        }

    def _chunk_text(self, text: str) -> list[dict]:
        """使用 LangChain Text Splitter 分块"""
        try:
            from langchain.text_splitter import RecursiveCharacterTextSplitter

            splitter = RecursiveCharacterTextSplitter(
                chunk_size=settings.chunk_size,
                chunk_overlap=settings.chunk_overlap,
                length_function=len,
                separators=["\n\n", "\n", "。", ".", "!", "！", "?", "？", " "],
            )
            chunks = splitter.split_text(text)
            return [
                {"index": i, "content": chunk}
                for i, chunk in enumerate(chunks)
            ]
        except Exception as e:
            logger.warning("LangChain 分块失败: %s", e)
            # fallback 到简单分块
            return self._chunk_text_simple(text)

    def _chunk_text_simple(self, text: str) -> list[dict]:
        """简单分块（fallback）"""
        chunks = []
        chunk_size = settings.chunk_size
        chunk_overlap = settings.chunk_overlap

        start = 0
        index = 0
        while start < len(text):
            end = start + chunk_size
            chunk = text[start:end]
            if chunk.strip():
                chunks.append({"index": index, "content": chunk})
                index += 1
            start = end - chunk_overlap

        return chunks

    async def _compute_embeddings(self, chunks: list[dict]) -> list[list[float] | None]:
        """计算嵌入向量（混合方案）"""
        embeddings = []
        for i, chunk in enumerate(chunks):
            try:
                # 优先使用本地模型
                embedding = await self._get_local_embedding(chunk["content"])
                if embedding is None:
                    # fallback 到 API
                    embedding = await self._get_api_embedding(chunk["content"])
                embeddings.append(embedding)
            except Exception as e:
                logger.warning("嵌入计算失败 chunk %d: %s", i, e)
                embeddings.append(None)
        return embeddings

    async def _get_local_embedding(self, text: str) -> list[float] | None:
        """使用本地模型计算嵌入（待实现）"""
        # TODO: 集成本地嵌入模型（BGE/Jina）
        return None

    async def _get_api_embedding(self, text: str) -> list[float] | None:
        """使用 API 计算嵌入"""
        try:
            resp = await self._gateway.embed(text, settings.embedding_model)
            return resp.embedding
        except Exception as e:
            logger.warning("API 嵌入计算失败: %s", e)
            return None

    async def _store_vectors(
        self,
        kb_id: str,
        doc_id: str,
        tenant_id: str,
        chunks: list[dict],
        embeddings: list[list[float] | None],
        db_type: VectorDBType,
    ):
        """存储向量"""
        if db_type == VectorDBType.MILVUS:
            await self._store_milvus(kb_id, doc_id, tenant_id, chunks, embeddings)
        elif db_type == VectorDBType.PGVECTOR:
            await self._store_pgvector(kb_id, doc_id, tenant_id, chunks, embeddings)
        else:
            raise ValueError(f"不支持的向量数据库: {db_type}")

    async def _store_milvus(self, kb_id, doc_id, tenant_id, chunks, embeddings):
        """存储到 Milvus"""
        if self._vector_store:
            # 使用 VectorStore 接口
            await self._store_with_interface(kb_id, doc_id, tenant_id, chunks, embeddings)
        else:
            # 直接使用 pymilvus
            await self._store_milvus_direct(kb_id, doc_id, tenant_id, chunks, embeddings)

    async def _store_with_interface(self, kb_id, doc_id, tenant_id, chunks, embeddings):
        """使用 VectorStore 接口存储"""
        collection = f"kb_{kb_id.replace('-', '_')}"
        dim = len(embeddings[0]) if embeddings and embeddings[0] else settings.embedding_dim

        await self._vector_store.ensure_collection(collection, dim)

        ids = []
        vectors = []
        payloads = []
        for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
            if embedding is None:
                continue
            ids.append(str(uuid.uuid4()))
            vectors.append(embedding)
            payloads.append({
                "content": chunk["content"][:65000],
                "kb_id": kb_id,
                "doc_id": doc_id,
                "tenant_id": tenant_id,
                "chunk_index": chunk["index"],
            })

        if ids:
            await self._vector_store.insert(collection, ids, vectors, payloads)
        logger.info("VectorStore 存储完成: %d chunks", len(ids))

    async def _store_milvus_direct(self, kb_id, doc_id, tenant_id, chunks, embeddings):
        """直接存储到 Milvus"""
        from pymilvus import Collection, FieldSchema, CollectionSchema, DataType

        self._ensure_milvus()
        dim = len(embeddings[0]) if embeddings and embeddings[0] else settings.embedding_dim

        fields = [
            FieldSchema(name="id", dtype=DataType.VARCHAR, is_primary=True, max_length=64),
            FieldSchema(name="kb_id", dtype=DataType.VARCHAR, max_length=64),
            FieldSchema(name="doc_id", dtype=DataType.VARCHAR, max_length=64),
            FieldSchema(name="tenant_id", dtype=DataType.VARCHAR, max_length=64),
            FieldSchema(name="chunk_index", dtype=DataType.INT64),
            FieldSchema(name="content", dtype=DataType.VARCHAR, max_length=65535),
            FieldSchema(name="embedding", dtype=DataType.FLOAT_VECTOR, dim=dim),
        ]
        schema = CollectionSchema(fields, description="RAG knowledge chunks")
        collection_name = f"kb_{kb_id.replace('-', '_')}"
        try:
            collection = Collection(collection_name)
        except Exception:
            collection = Collection(collection_name, schema)
            collection.create_index("embedding", {
                "metric_type": "COSINE",
                "index_type": "IVF_FLAT",
                "params": {"nlist": 1024},
            })

        ids, kb_ids, doc_ids, tenant_ids, chunk_indices, contents, valid_embeddings = [], [], [], [], [], [], []
        for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
            if embedding is None:
                continue
            ids.append(str(uuid.uuid4()))
            kb_ids.append(kb_id)
            doc_ids.append(doc_id)
            tenant_ids.append(tenant_id)
            chunk_indices.append(chunk["index"])
            contents.append(chunk["content"][:65000])
            valid_embeddings.append(embedding)

        if ids:
            await asyncio.to_thread(collection.insert, [ids, kb_ids, doc_ids, tenant_ids, chunk_indices, contents, valid_embeddings])
            await asyncio.to_thread(collection.flush)
        logger.info("Milvus 存储完成: %d chunks", len(ids))

    async def _store_pgvector(self, kb_id, doc_id, tenant_id, chunks, embeddings):
        """存储到 PostgreSQL (pgvector)"""
        # TODO: 实现 pgvector 存储
        raise NotImplementedError("pgvector 存储尚未实现")

    async def query(
        self,
        kb_id: str,
        query: str,
        top_k: int = 5,
        threshold: float = 0.5,
        vector_db: str | None = None,
    ) -> list[dict]:
        """查询知识库"""
        db_type = VectorDBType(vector_db) if vector_db else self._vector_db_type

        if db_type == VectorDBType.MILVUS:
            return await self.query_milvus(kb_id, query, top_k, threshold)
        elif db_type == VectorDBType.PGVECTOR:
            return await self.query_pgvector(kb_id, query, top_k, threshold)
        else:
            raise ValueError(f"不支持的向量数据库: {db_type}")

    async def query_milvus(self, kb_id: str, query: str, top_k: int = 5, threshold: float = 0.5) -> list[dict]:
        """从 Milvus 查询"""
        if self._vector_store:
            return await self._query_with_interface(kb_id, query, top_k, threshold)

        from pymilvus import Collection
        self._ensure_milvus()
        collection_name = f"kb_{kb_id.replace('-', '_')}"
        try:
            collection = Collection(collection_name)
            collection.load()
        except Exception as e:
            logger.error("Milvus collection load failed: %s", e)
            return []

        resp = await self._gateway.embed(query, settings.embedding_model)
        if not resp.embedding:
            return []

        results = await asyncio.to_thread(
            collection.search,
            data=[resp.embedding], anns_field="embedding",
            param={"metric_type": "COSINE", "params": {"nprobe": 10}},
            limit=top_k, output_fields=["doc_id", "chunk_index", "content"],
        )
        return [
            {"id": hit.id, "content": hit.entity.get("content", ""),
             "doc_id": hit.entity.get("doc_id", ""), "chunk_index": hit.entity.get("chunk_index", 0),
             "score": hit.score}
            for hits in results for hit in hits if hit.score >= threshold
        ]

    async def _query_with_interface(self, kb_id, query, top_k, threshold) -> list[dict]:
        """使用 VectorStore 接口查询"""
        collection = f"kb_{kb_id.replace('-', '_')}"

        resp = await self._gateway.embed(query, settings.embedding_model)
        if not resp.embedding:
            return []

        results = await self._vector_store.search(
            collection=collection,
            query_vector=resp.embedding,
            top_k=top_k,
            threshold=threshold,
        )

        return [
            {
                "id": r["id"],
                "content": r.get("content", ""),
                "doc_id": r.get("doc_id", ""),
                "chunk_index": r.get("chunk_index", 0),
                "score": r.get("score", 0),
            }
            for r in results
        ]

    async def query_pgvector(self, kb_id: str, query: str, top_k: int = 5, threshold: float = 0.5) -> list[dict]:
        """从 PostgreSQL (pgvector) 查询"""
        # TODO: 实现 pgvector 查询
        raise NotImplementedError("pgvector 查询尚未实现")

    async def query_qdrant(self, kb_id: str, query: str, top_k: int = 5, threshold: float = 0.5) -> list[dict]:
        """从 Qdrant 查询（兼容旧代码）"""
        from qdrant_client import QdrantClient
        client = QdrantClient(host="localhost", port=6333)
        collection_name = f"kb_{kb_id.replace('-', '_')}"
        resp = await self._gateway.embed(query, settings.embedding_model)
        if not resp.embedding:
            return []
        results = await asyncio.to_thread(
            client.search,
            collection_name=collection_name, query_vector=resp.embedding,
            limit=top_k, score_threshold=threshold,
        )
        return [{"id": hit.id, "content": hit.payload.get("content", ""),
                 "doc_id": hit.payload.get("doc_id", ""), "chunk_index": hit.payload.get("chunk_index", 0),
                 "score": hit.score} for hit in results]
